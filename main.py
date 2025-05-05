
import os
import pandas as pd
from jira import JIRA
from dotenv import load_dotenv
from datetime import datetime
import yagmail
import argparse
import openpyxl
from openpyxl.styles import PatternFill

load_dotenv()

def get_missing_fields(issue):
    missing = []
    if not issue.fields.summary:
        missing.append("Summary")
    if not issue.fields.description:
        missing.append("Description")
    if not issue.fields.fixVersions:
        missing.append("Fix Version")
    if not getattr(issue.fields, "customfield_10008", None):  # Epic
        missing.append("Epic")
    if not getattr(issue.fields, "customfield_10011", None):  # Acceptance Criteria
        missing.append("Acceptance Criteria")
    return missing

def fetch_invalid_issues(project, fix_version):
    options = {"server": os.getenv("JIRA_URL")}
    jira = JIRA(options, basic_auth=(os.getenv("JIRA_USER"), os.getenv("JIRA_PASS")))
    jql = f'project = {project} AND status NOT IN (Open, Closed)'
    if fix_version:
        jql += f' AND fixVersion = "{fix_version}"'
    issues = jira.search_issues(jql, maxResults=1000)
    invalid_data = []
    for issue in issues:
        missing = get_missing_fields(issue)
        if missing:
            invalid_data.append({
                "Issue Key": issue.key,
                "Summary": issue.fields.summary,
                "Status": issue.fields.status.name,
                "Missing Fields": ", ".join(missing)
            })
    return invalid_data

def generate_excel_report(data, file_path):
    df = pd.DataFrame(data)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Invalid JIRAs')
        ws = writer.sheets['Invalid JIRAs']
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.fill = red_fill

def build_html(project, fix_version, data):
    rows = ""
    for item in data:
        rows += f"<tr><td>{item['Issue Key']}</td><td>{item['Summary']}</td><td>{item['Status']}</td><td>{item['Missing Fields']}</td></tr>"
    return f"""
    <html>
    <body>
        <h2>JIRA Missing Fields Report</h2>
        <p><strong>Project:</strong> {project}</p>
        <p><strong>Fix Version:</strong> {fix_version or "N/A"}</p>
        <p><strong>Number of Invalid JIRAs:</strong> {len(data)}</p>
        <table border="1" cellpadding="5" cellspacing="0">
            <tr>
                <th>Issue Key</th><th>Summary</th><th>Status</th><th>Missing Fields</th>
            </tr>
            {rows}
        </table>
    </body>
    </html>
    """

def send_email(recipients, subject, body, attachment):
    yag = yagmail.SMTP(user=os.getenv("SMTP_USER"), password=os.getenv("SMTP_PASS"))
    yag.send(to=recipients, subject=subject, contents=[body, attachment])

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", required=True)
    parser.add_argument("--fix-version", required=False)
    parser.add_argument("--mail-group", required=True)
    args = parser.parse_args()

    data = fetch_invalid_issues(args.project, args.fix_version)
    if not data:
        print("No invalid JIRA issues found.")
        exit(0)

    filename = f"jira_missing_fields_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    generate_excel_report(data, filename)

    html_body = build_html(args.project, args.fix_version, data)
    recipients = args.mail_group.split(",")

    send_email(recipients, f"[JIRA] Missing Fields Report - {args.project}", html_body, filename)
    print("Report sent successfully.")
